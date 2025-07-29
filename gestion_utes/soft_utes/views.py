import json
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Subquery
from .pdf_utils import generar_pdf
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.db import IntegrityError
from django.shortcuts import render, get_object_or_404, redirect
from .models import (
    ActaSalida, Area, AsignacionEquipo, Cargo, Establecimiento, Formato, Lote, Personal, SubArea, TipoEquipo, Marca, Modelo, EquiposInformatico, 
    Componentes, Caracteristicas, ComponenteCaracteristica, TipoComponente
)
from django.http import HttpResponse, HttpResponseBadRequest, JsonResponse
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import make_password
from datetime import datetime



@login_required
def logout_view(request):
    logout(request)
    messages.success(request, "Sesión cerrada correctamente.")
    return redirect('login')


def login_view(request):
    error_login = None  # Inicializa la variable de error

    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '').strip()

        if not email or not password:
            error_login = 'Todos los campos son obligatorios.'
        else:
            try:
                user = User.objects.get(email=email)
                user = authenticate(request, username=user.username, password=password)

                if user is not None:
                    login(request, user)

                    if user.is_superuser:
                        return redirect('dashboard_admin')
                    else:
                        return redirect('dashboard_gestor')
                else:
                    error_login = 'Contraseña incorrecta.'
            except User.DoesNotExist:
                error_login = 'No existe un usuario con este correo electrónico.'

    return render(request, 'login.html', {'error_login': error_login})

def restablecer_contrasena(request):
    if request.method == 'POST':
        correo = request.POST.get('usuario')
        nueva = request.POST.get('nueva_contrasena')
        confirmar = request.POST.get('confirmar_contrasena')

        if nueva != confirmar:
            messages.error(request, "Las contraseñas no coinciden.")
            return redirect('login')  

        try:
            user = User.objects.get(email=correo)
            user.set_password(nueva)
            user.save()
            messages.success(request, "Contraseña actualizada correctamente.")
        except User.DoesNotExist:
            messages.error(request, "No se encontró un usuario con ese correo.")
        
        return redirect('login')  
    else:
        return redirect('login')

@login_required
def dashboard_admin(request):
    if not request.user.is_superuser:
        messages.error(request, 'Acceso no autorizado')
        return redirect('login')
    
    # Configuración de paginación
    items_per_page = 10
    
    # Procesamiento de equipos
    equipos_list = EquiposInformatico.objects.all().order_by('-idEquipoInformatico')
    equipos_page = request.GET.get('equipos_page')
    equipos_paginator = Paginator(equipos_list, items_per_page)
    
    # Procesamiento de componentes
    componentes_list = Componentes.objects.all().order_by('-idComponente')
    componentes_page = request.GET.get('componentes_page')
    componentes_paginator = Paginator(componentes_list, items_per_page)
    
    # Procesamiento de asignaciones
    asignaciones_list = AsignacionEquipo.objects.select_related(
        'idEquipoInformatico', 'idActaSalida'
    ).order_by('-fechaAsignacion')
    asignaciones_page = request.GET.get('asignaciones_page')
    asignaciones_paginator = Paginator(asignaciones_list, items_per_page)
    
    try:
        equipos = equipos_paginator.page(equipos_page) if equipos_page else equipos_paginator.page(1)
    except (PageNotAnInteger, EmptyPage):
        equipos = equipos_paginator.page(1)
    
    try:
        componentes = componentes_paginator.page(componentes_page) if componentes_page else componentes_paginator.page(1)
    except (PageNotAnInteger, EmptyPage):
        componentes = componentes_paginator.page(1)
    
    try:
        asignaciones = asignaciones_paginator.page(asignaciones_page) if asignaciones_page else asignaciones_paginator.page(1)
    except (PageNotAnInteger, EmptyPage):
        asignaciones = asignaciones_paginator.page(1)
    
    context = {
        'equipos_count': equipos_list.count(),
        'componentes_count': componentes_list.count(),
        'asignaciones_count': asignaciones_list.count(),
        'equipos_page': equipos,
        'componentes_page': componentes,
        'asignaciones_page': asignaciones,
    }
    
    return render(request, 'admin_dashboard.html', context)

@login_required
def dashboard_gestor(request):
    if request.user.is_superuser:
        return redirect('dashboard_admin')
    
# Configuración de paginación
    items_per_page = 10
    
    # Procesamiento de equipos
    equipos_list = EquiposInformatico.objects.all().order_by('-idEquipoInformatico')
    equipos_page = request.GET.get('equipos_page')
    equipos_paginator = Paginator(equipos_list, items_per_page)
    
    # Procesamiento de componentes
    componentes_list = Componentes.objects.all().order_by('-idComponente')
    componentes_page = request.GET.get('componentes_page')
    componentes_paginator = Paginator(componentes_list, items_per_page)
    
    # Procesamiento de asignaciones
    asignaciones_list = AsignacionEquipo.objects.select_related(
        'idEquipoInformatico', 'idActaSalida'
    ).order_by('-fechaAsignacion')
    asignaciones_page = request.GET.get('asignaciones_page')
    asignaciones_paginator = Paginator(asignaciones_list, items_per_page)
    
    try:
        equipos = equipos_paginator.page(equipos_page) if equipos_page else equipos_paginator.page(1)
    except (PageNotAnInteger, EmptyPage):
        equipos = equipos_paginator.page(1)
    
    try:
        componentes = componentes_paginator.page(componentes_page) if componentes_page else componentes_paginator.page(1)
    except (PageNotAnInteger, EmptyPage):
        componentes = componentes_paginator.page(1)
    
    try:
        asignaciones = asignaciones_paginator.page(asignaciones_page) if asignaciones_page else asignaciones_paginator.page(1)
    except (PageNotAnInteger, EmptyPage):
        asignaciones = asignaciones_paginator.page(1)
    
    context = {
        'equipos_count': equipos_list.count(),
        'componentes_count': componentes_list.count(),
        'asignaciones_count': asignaciones_list.count(),
        'equipos_page': equipos,
        'componentes_page': componentes,
        'asignaciones_page': asignaciones,
    }
    return render(request, 'gestor_dashboard.html', context)

# ---------------------- TIPO EQUIPO ----------------------
def tipoequipo_list(request):
    tipos = TipoEquipo.objects.all()
    return render(request, 'tipoequipo_list.html', {'tipos': tipos})

def tipoequipo_create(request):
    if request.method == 'POST':
        nombre = request.POST['nombre']
        estado = request.POST.get('estado', 'ACTIVO')
        TipoEquipo.objects.create(nombre=nombre, estado=estado)
        return redirect('tipoequipo_list')
    return render(request, 'tipoequipo_list.html')

def tipoequipo_update(request, idTipoEquipo):
    tipo = get_object_or_404(TipoEquipo, idTipoEquipo=idTipoEquipo)
    if request.method == 'POST':
        tipo.nombre = request.POST['nombre']
        tipo.estado = request.POST.get('estado', 'ACTIVO')
        tipo.save()
        return redirect('tipoequipo_list')
    return render(request, 'tipoequipo_list.html', {'tipo': tipo})

def tipoequipo_delete(request, idTipoEquipo):
    tipo = get_object_or_404(TipoEquipo, idTipoEquipo=idTipoEquipo)
    tipo.estado = "INACTIVO"
    tipo.save()
    return redirect('tipoequipo_list')

def reactivar_tipoequipo(request, idTipoEquipo):
    tipo = get_object_or_404(TipoEquipo, idTipoEquipo=idTipoEquipo)
    tipo.estado = "ACTIVO"
    tipo.save()
    return redirect('tipoequipo_list')

# ---------------------- MARCA ----------------------
def marca_list(request):
    marcas = Marca.objects.all()
    return render(request, 'marca_list.html', {'marcas': marcas})

def marca_create(request):
    if request.method == 'POST':
        nombre = request.POST['nombreMarca']
        Marca.objects.create(nombreMarca=nombre)
        return redirect('marca_list')
    return render(request, 'marca_list.html')

def marca_update(request, idMarca):
    marca = get_object_or_404(Marca, idMarca=idMarca)
    if request.method == 'POST':
        marca.nombreMarca = request.POST['nombreMarca']
        marca.save()
        return redirect('marca_list')
    return render(request, 'marca_list.html', {'marca': marca})

def marca_delete(request, idMarca):
    marca = get_object_or_404(Marca, idMarca=idMarca)
    marca.estado = "INACTIVO"
    marca.save()
    return redirect('marca_list')

def reactivar_marca(request, id):
    marca = get_object_or_404(Marca, idMarca=id)
    marca.estado = "ACTIVO"
    marca.save()
    return redirect('marca_list')

# ---------------------- MODELO ----------------------
def modelo_list(request):
    modelos = Modelo.objects.select_related('idMarca').all()
    marcas = Marca.objects.filter(estado="ACTIVO")
    return render(request, 'modelo_list.html', {
        'modelos': modelos,
        'marcas': marcas
    })

def modelo_create(request):
    if request.method == 'POST':
        marca = get_object_or_404(Marca, idMarca=request.POST['idMarca'])
        modelo = Modelo(
            idMarca=marca,
            nombreModelo=request.POST['nombreModelo'],
            descripcionModelo=request.POST.get('descripcionModelo', '')
        )
        if 'Imagen' in request.FILES:
            modelo.Imagen = request.FILES['Imagen']
        modelo.save()
        return redirect('modelo_list')
    
    marcas = Marca.objects.all()
    return render(request, 'modelo_list.html', {'marcas': marcas})

def modelo_update(request, idModelo):
    modelo = get_object_or_404(Modelo, idModelo=idModelo)
    if request.method == 'POST':
        modelo.idMarca = get_object_or_404(Marca, idMarca=request.POST['idMarca'])
        modelo.nombreModelo = request.POST['nombreModelo']
        modelo.descripcionModelo = request.POST.get('descripcionModelo', '')
        if 'Imagen' in request.FILES:
            modelo.Imagen = request.FILES['Imagen']
        modelo.save()
        return redirect('modelo_list')
    
    marcas = Marca.objects.all()
    return render(request, 'modelo_list.html', {
        'modelo': modelo,
        'marcas': marcas
    })

def modelo_delete(request, idModelo):
    modelo = get_object_or_404(Modelo, idModelo=idModelo)
    modelo.estado = "INACTIVO"
    modelo.save()
    return redirect('modelo_list')

def reactivar_modelo(request, id):
    modelo = get_object_or_404(Modelo, idModelo=id)
    modelo.estado = "ACTIVO"
    modelo.save()
    return redirect('modelo_list')

# ---------------------- EQUIPOS INFORMATICOS ----------------------
def equipo_list(request):
    equipos = EquiposInformatico.objects.select_related('idTipoEquipo', 'idModelo').all()
    tipos = TipoEquipo.objects.all()
    modelos = Modelo.objects.all()
    return render(request, 'equipo_list.html', {
        'equipos': equipos,
        'tipos': tipos,
        'modelos': modelos
    })

def equipo_detalle(request, idEquipoInformatico):
    equipo = get_object_or_404(EquiposInformatico, idEquipoInformatico=idEquipoInformatico)
    componentes = Componentes.objects.filter(idEquipoInformatico=equipo)
    return render(request, 'detalle_equipo.html', {
        'equipo': equipo,
        'componentes': componentes
    })
def asignar_componentes(request, idEquipoInformatico):
    equipo = get_object_or_404(EquiposInformatico, idEquipoInformatico=idEquipoInformatico)

    if request.method == 'POST':
        componente_id = request.POST.get('componente_id')
        accion = request.POST.get('accion')

        if componente_id and accion:
            componente = get_object_or_404(Componentes, idComponente=componente_id)

            if accion == 'asignar':
                componente.idEquipoInformatico = equipo
                componente.estado = 'ASIGNADO'
            elif accion == 'quitar':
                componente.idEquipoInformatico = None
                componente.estado = 'ACTIVO'

            componente.save()

    buscar_asignados = request.GET.get('buscar_asignados', '')
    buscar_disponibles = request.GET.get('buscar_disponibles', '')

    q_asignados = Q(numeroSerie__icontains=buscar_asignados) | Q(descripcionComponente__icontains=buscar_asignados) | Q(idModelo__nombreModelo__icontains=buscar_asignados) | Q(idModelo__idMarca__nombreMarca__icontains=buscar_asignados)
    q_disponibles = Q(numeroSerie__icontains=buscar_disponibles) | Q(idTipoComponente__nombre__icontains=buscar_disponibles) | Q(descripcionComponente__icontains=buscar_disponibles) | Q(idModelo__nombreModelo__icontains=buscar_disponibles) | Q(idModelo__idMarca__nombreMarca__icontains=buscar_disponibles)

    componentes_asignados = Componentes.objects.filter(
        idEquipoInformatico=equipo,
        estado='ASIGNADO'
    ).filter(q_asignados)

    componentes_disponibles = Componentes.objects.filter(
        estado='ACTIVO'
    ).filter(q_disponibles)

    paginator_asignados = Paginator(componentes_asignados, 10)
    paginator_disponibles = Paginator(componentes_disponibles, 10)

    pagina_asignados = request.GET.get('pagina_asignados')
    pagina_disponibles = request.GET.get('pagina_disponibles')

    componentes_asignados = paginator_asignados.get_page(pagina_asignados)
    componentes_disponibles = paginator_disponibles.get_page(pagina_disponibles)

    return render(request, 'asignacion_componentes.html', {
        'equipo': equipo,
        'componentes_disponibles': componentes_disponibles,
        'componentes_asignados': componentes_asignados,
        'buscar_asignados': buscar_asignados,
        'buscar_disponibles': buscar_disponibles
    })

def equipo_create(request):
    if request.method == 'POST':
        try:
            tipo_id = request.POST.get('idTipoEquipo')
            if not tipo_id:
                return JsonResponse({"error": "Debe seleccionar un tipo de equipo."}, status=400)

            tipo = get_object_or_404(TipoEquipo, idTipoEquipo=tipo_id)
            modelo = get_object_or_404(Modelo, idModelo=request.POST['idModelo']) if request.POST.get('idModelo') else None
            lote = get_object_or_404(Lote, idLote=request.POST['idLote']) if request.POST.get('idLote') else None

            equipo = EquiposInformatico(
                nombreEquipoInformatico=request.POST['nombreEquipoInformatico'],
                idTipoEquipo=tipo,
                idModelo=modelo,
                idLote=lote,
                codigoPatrimonial=request.POST.get('codigoPatrimonial'),
                numeroSerie=request.POST.get('numeroSerie'),
                observacionEquipo=request.POST.get('observacionEquipo', ''),
                descripcionEquipo=request.POST.get('descripcionEquipo'),
                añoGarantia=request.POST.get('añoGarantia') or None,
                estado='ACTIVO'
            )
            equipo.save()

            return JsonResponse({"mensaje": "Equipo creado correctamente", "idEquipo": equipo.idEquipoInformatico})

        except IntegrityError:
            return JsonResponse({"error": "Código patrimonial o número de serie duplicado."}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error al crear equipo: {str(e)}"}, status=500)

    tipos = TipoEquipo.objects.all()
    modelos = Modelo.objects.all()
    return render(request, 'equipo_list.html', {
        'tipos': tipos,
        'modelos': modelos,
        'descripciones': EquiposInformatico.Tipo_descripcion
    })

def equipo_update(request, idEquipoInformatico):
    equipo = get_object_or_404(EquiposInformatico, pk=idEquipoInformatico)  # usa pk o asegúrate del nombre exacto

    if request.method == 'POST':
        try:
            codigoPatrimonial = request.POST.get('codigoPatrimonial')
            numeroSerie = request.POST.get('numeroSerie')

            duplicado = EquiposInformatico.objects.filter(
                Q(codigoPatrimonial=codigoPatrimonial) | Q(numeroSerie=numeroSerie)
            ).exclude(pk=idEquipoInformatico).exists()

            if duplicado:
                messages.error(request, '⚠️ Ya existe un equipo con el mismo número de serie o código patrimonial.')
                raise ValueError("Duplicado detectado")

            nombre = request.POST.get('nombreEquipoInformatico')
            tipo = request.POST.get('idTipoEquipo')

            if not nombre or not tipo:
                messages.error(request, '❌ Faltan datos obligatorios.')
                return HttpResponseBadRequest("Faltan campos requeridos.")

            equipo.nombreEquipoInformatico = nombre
            equipo.idTipoEquipo = get_object_or_404(TipoEquipo, pk=tipo)

            modelo_id = request.POST.get('idModelo')
            equipo.idModelo = get_object_or_404(Modelo, pk=modelo_id) if modelo_id else None

            lote_id = request.POST.get('idLote')
            equipo.idLote = get_object_or_404(Lote, pk=lote_id) if lote_id else None

            equipo.codigoPatrimonial = codigoPatrimonial
            equipo.numeroSerie = numeroSerie
            equipo.observacionEquipo = request.POST.get('observacionEquipo', '')
            equipo.descripcionEquipo = request.POST.get('descripcionEquipo', '')
            equipo.añoGarantia = request.POST.get('añoGarantia') or None

            estado = request.POST.get('estado')
            if estado:
                equipo.estado = estado

            equipo.save()
            messages.success(request, '✅ Equipo actualizado exitosamente')
            return redirect('equipo_list')

        except Exception as e:
            print("❌ Error:", e)
            return HttpResponseBadRequest(f"Error: {e}")


def equipo_delete(request, idEquipoInformatico):
    equipo = get_object_or_404(EquiposInformatico, idEquipoInformatico=idEquipoInformatico)
    equipo.estado = 'INACTIVO'
    equipo.save()
    return redirect('equipo_list')

def reactivar_equipo(request, idEquipoInformatico):
    equipo = get_object_or_404(EquiposInformatico, idEquipoInformatico=idEquipoInformatico)
    equipo.estado = 'ACTIVO'
    equipo.save()
    return redirect('equipo_list')

# ---------------------- TIPO COMPONENTE ----------------------
def tipocomponente_list(request):
    tipos = TipoComponente.objects.all()
    return render(request, 'tipocomponente_list.html', {'tipos': tipos})

def tipocomponente_create(request):
    if request.method == 'POST':
        nombre = request.POST['nombre']
        estado = request.POST.get('estado', 'ACTIVO')
        TipoComponente.objects.create(nombre=nombre, estado=estado)
        return redirect('tipocomponente_list')
    return render(request, 'tipocomponente_list.html')

def tipocomponente_update(request, idTipoComponente):
    tipo = get_object_or_404(TipoComponente, idTipoComponente=idTipoComponente)
    if request.method == 'POST':
        tipo.nombre = request.POST['nombre']
        tipo.estado = request.POST.get('estado', 'ACTIVO')
        tipo.save()
        return redirect('tipocomponente_list')
    return render(request, 'tipocomponente_list.html', {'tipo': tipo})

def tipocomponente_delete(request, idTipoComponente):
    tipo = get_object_or_404(TipoComponente, idTipoComponente=idTipoComponente)
    tipo.estado = "INACTIVO"
    tipo.save()
    return redirect('tipocomponente_list')

def reactivar_tipocomponente(request, idTipoComponente):
    tipo = get_object_or_404(TipoComponente, idTipoComponente=idTipoComponente)
    tipo.estado = "ACTIVO"
    tipo.save()
    return redirect('tipocomponente_list')

# ---------------------- CARACTERISTICAS ----------------------
def caracteristica_list(request):
    caracteristicas = Caracteristicas.objects.select_related('idTipoComponente').all()
    tipos = TipoComponente.objects.all()
    return render(request, 'caracteristica_list.html', {
        'caracteristicas': caracteristicas,
        'tipos': tipos
    })

def caracteristica_create(request):
    if request.method == 'POST':
        try:
            tipo = get_object_or_404(TipoComponente, idTipoComponente=request.POST['idTipoComponente'])
            Caracteristicas.objects.create(
                descripcionCaracteristica=request.POST['descripcionCaracteristica'],
                idTipoComponente=tipo
            )
            messages.success(request, 'Característica creada exitosamente')
            return redirect('caracteristica_list')
        except Exception as e:
            messages.error(request, f'Error al crear característica: {str(e)}')
    
    tipos = TipoComponente.objects.all()
    return render(request, 'caracteristica_list.html', {'tipos': tipos})

def caracteristica_update(request, idCaracteristica):
    caracteristica = get_object_or_404(Caracteristicas, idCaracteristica=idCaracteristica)
    if request.method == 'POST':
        try:
            tipo = get_object_or_404(TipoComponente, idTipoComponente=request.POST['idTipoComponente'])
            caracteristica.descripcionCaracteristica = request.POST['descripcionCaracteristica']
            caracteristica.idTipoComponente = tipo
            caracteristica.save()
            messages.success(request, 'Característica actualizada exitosamente')
            return redirect('caracteristica_list')
        except Exception as e:
            messages.error(request, f'Error al actualizar característica: {str(e)}')
    
    tipos = TipoComponente.objects.all()
    return render(request, 'caracteristica_list.html', {
        'caracteristica': caracteristica,
        'tipos': tipos
    })

def caracteristica_delete(request, idCaracteristica):
    caracteristica = get_object_or_404(Caracteristicas, idCaracteristica=idCaracteristica)
    caracteristica.estado = 'INACTIVO'
    caracteristica.save()
    messages.success(request, 'Característica eliminada exitosamente')
    return redirect('caracteristica_list')

def reactivar_caracteristica(request, idCaracteristica):
    caracteristica = get_object_or_404(Caracteristicas, idCaracteristica=idCaracteristica)
    caracteristica.estado = 'ACTIVO'
    caracteristica.save()
    messages.success(request, 'Característica eliminada exitosamente')
    return redirect('caracteristica_list')

# ---------------------- COMPONENTES ----------------------
def componente_list(request):
    tipos = TipoComponente.objects.filter(estado='ACTIVO')
    modelos = Modelo.objects.filter(estado='ACTIVO')

    id_tipo = request.GET.get('idTipoComponente')
    
    componentes = Componentes.objects.select_related('idModelo', 'idTipoComponente')

    if id_tipo:
        componentes = componentes.filter(idTipoComponente_id=id_tipo)

    # Agregar características como JSON a cada componente
    for comp in componentes:
        caracteristicas = ComponenteCaracteristica.objects.filter(idComponente=comp)
        comp.caracteristicas_json = json.dumps({
            f"caracteristica_{c.idCaracteristica.idCaracteristica}": c.valor for c in caracteristicas
        })

    return render(request, 'componente_list.html', {
        'componentes': componentes,
        'modelos': modelos,
        'tipos': tipos,
        'filtro_tipo': id_tipo,
        'descripciones': Componentes.Tipo_descripcion
    })

def componente_create(request):
    if request.method == 'POST':
        try:
            if not request.POST.get('idModelo'):
                raise ValueError("El campo modelo es requerido")
            if not request.POST.get('idTipoComponente'):
                raise ValueError("El tipo de componente es requerido")

            modelo = get_object_or_404(Modelo, idModelo=request.POST['idModelo'])
            tipo = get_object_or_404(TipoComponente, idTipoComponente=request.POST['idTipoComponente'])

            # Buscar o crear lote si viene desde formulario múltiple
            lote = None
            cantidad = request.POST.get('cantidad')
            fecha_adq = request.POST.get('fechaAdquisicion')

            if cantidad and fecha_adq:
                # Reutilizar lote si ya existe uno con mismo modelo, fecha y cantidad
                lote, creado = Lote.objects.get_or_create(
                    cantidad=int(cantidad),
                    idModelo=modelo,
                    fechaAdquisicion=fecha_adq
                )

            componente = Componentes(
                idModelo=modelo,
                idTipoComponente=tipo,
                numeroSerie=request.POST.get('numeroSerie', '').strip(),
                descripcionComponente=request.POST.get('descripcionComponente', 'De Marca'),
                estado='ACTIVO'
            )

            if lote:
                componente.idLote = lote  # solo si tienes este campo en tu modelo

            componente.save()

            # Características comunes
            for key, value in request.POST.items():
                if key.startswith('caracteristica_'):
                    try:
                        carac_id = key.split('_')[1]
                        caracteristica = get_object_or_404(Caracteristicas, idCaracteristica=carac_id)
                        ComponenteCaracteristica.objects.update_or_create(
                            idComponente=componente,
                            idCaracteristica=caracteristica,
                            defaults={'valor': value.strip()}
                        )
                    except Exception as e:
                        print(f"Error al procesar característica {key}: {str(e)}")

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'ok': "ACTIVO"})

            messages.success(request, 'Componente creado exitosamente')
            return redirect('componente_list')

        except Exception as e:
            mensaje = f'Error al crear componente: {str(e)}'
            print("❌", mensaje)
            print("📤 Datos recibidos:", request.POST)

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'error': mensaje}, status=500)

            messages.error(request, mensaje)

    modelos = Modelo.objects.all().select_related('idMarca')
    tipos = TipoComponente.objects.all()
    return render(request, 'componente_list.html', {
        'modelos': modelos,
        'tipos': tipos,
        'descripciones': Componentes.Tipo_descripcion
    })

def componente_update(request, idComponente):
    componente = get_object_or_404(Componentes, idComponente=idComponente)

    if request.method == 'POST':
        try:
            modelo = get_object_or_404(Modelo, idModelo=request.POST['idModelo'])
            tipo = get_object_or_404(TipoComponente, idTipoComponente=request.POST['idTipoComponente'])
            
            componente.idModelo = modelo
            componente.idTipoComponente = tipo
            componente.numeroSerie = request.POST.get('numeroSerie', '')
            componente.descripcionComponente = request.POST.get('descripcionComponente', 'De Marca')
         
            componente.save()

            # Actualizar o crear características
            for key, value in request.POST.items():
                if key.startswith('caracteristica_'):
                    carac_id = key.split('_')[1]
                    caracteristica = get_object_or_404(Caracteristicas, idCaracteristica=carac_id)
                    ComponenteCaracteristica.objects.update_or_create(
                        idComponente=componente,
                        idCaracteristica=caracteristica,
                        defaults={'valor': value}
                    )

            messages.success(request, 'Componente actualizado exitosamente')
            return redirect('componente_list')

        except Exception as e:
            messages.error(request, f'Error al actualizar componente: {str(e)}')

    modelos = Modelo.objects.all()
    tipos = TipoComponente.objects.all()

    return render(request, 'componente_list.html', {
        'modelos': modelos,
        'tipos': tipos,
        'descripciones': Componentes.Tipo_descripcion,
        'editando': "ACTIVO",
        'componente': componente,
    })

def componente_delete(request, idComponente):
    componente = get_object_or_404(Componentes, idComponente=idComponente)
    componente.estado = 'INACTIVO'
    componente.save()
    messages.success(request, 'Componente dado de baja exitosamente')
    return redirect('componente_list')

def reactivar_componente(request, idComponente):
    componente = get_object_or_404(Componentes, idComponente=idComponente)
    componente.estado = 'ACTIVO'
    componente.save()
    return redirect('componente_list')

# ---------------------- COMPONENTE-CARACTERISTICA ----------------------
def comp_carac_list(request):
    relaciones = ComponenteCaracteristica.objects.select_related(
        'idComponente', 'idCaracteristica'
    ).all()
    return render(request, 'compcarac_list.html', {'relaciones': relaciones})

def comp_carac_create(request):
    if request.method == 'POST':
        try:
            componente = get_object_or_404(Componentes, idComponente=request.POST.get('idComponente'))
            caracteristica = get_object_or_404(Caracteristicas, idCaracteristica=request.POST.get('idCaracteristica'))
            
            ComponenteCaracteristica.objects.create(
                idComponente=componente,
                idCaracteristica=caracteristica,
                valor=request.POST.get('valor', '')
            )
            messages.success(request, 'Relación componente-característica creada exitosamente')
            return redirect('comp_carac_list')
        except Exception as e:
            messages.error(request, f'Error al crear la relación: {str(e)}')
    
    componentes = Componentes.objects.all()
    caracteristicas = Caracteristicas.objects.all()
    return render(request, 'compcarac_list.html', {
        'componentes': componentes,
        'caracteristicas': caracteristicas
    })

def comp_carac_update(request, idComponenteCaracteristica):
    relacion = get_object_or_404(ComponenteCaracteristica, idComponenteCaracteristica=idComponenteCaracteristica)
    if request.method == 'POST':
        relacion.valor = request.POST['valor']
        relacion.save()
        messages.success(request, 'Valor actualizado exitosamente')
        return redirect('comp_carac_list')
    
    return render(request, 'compcarac_list.html', {'relacion': relacion})

def comp_carac_delete(request, idComponenteCaracteristica):
    relacion = get_object_or_404(ComponenteCaracteristica, idComponenteCaracteristica=idComponenteCaracteristica)
    relacion.delete()
    messages.success(request, 'Relación eliminada exitosamente')
    return redirect('comp_carac_list')

def componente_detalle(request, idComponente):
    componente = get_object_or_404(Componentes, pk=idComponente)
    caracteristicas = ComponenteCaracteristica.objects.filter(idComponente=componente)
    context = {
        'componente': componente,
        'caracteristicas': caracteristicas,
    }
    return render(request, 'detalle_componente.html', context)
# ---------------------- AJAX ----------------------
def caracteristicas_por_tipo(request, tipo_id):
    caracteristicas = Caracteristicas.objects.filter(
        idTipoComponente=tipo_id,
        estado='ACTIVO'
    ).values('idCaracteristica', 'descripcionCaracteristica')
    return JsonResponse(list(caracteristicas), safe=False)

# ---------------------- ESTABLECIMIENTO ----------------------
# CREATE
def crear_establecimiento(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        Establecimiento.objects.create(NombreEstablecimiento=nombre)
        return redirect('listar_establecimientos')
    return render(request, 'establecimiento_list.html')

# READ
def listar_establecimientos(request):
    establecimientos = Establecimiento.objects.all()
    return render(request, 'establecimiento_list.html', {'establecimientos': establecimientos})

# UPDATE
def editar_establecimiento(request, id):
    establecimiento = get_object_or_404(Establecimiento, idEstablecimiento=id)
    if request.method == 'POST':
        establecimiento.NombreEstablecimiento = request.POST.get('nombre')
        establecimiento.save()
        return redirect('listar_establecimientos')
    return render(request, 'establecimiento_list.html', {'establecimiento': establecimiento})

# DELETE
def eliminar_establecimiento(request, id):
    establecimiento = get_object_or_404(Establecimiento, idEstablecimiento=id)
    establecimiento.estado = "INACTIVO"
    establecimiento.save()
    return redirect('listar_establecimientos')

# REACTIVATE

def reactivar_establecimiento(request, id):
    establecimiento = get_object_or_404(Establecimiento, idEstablecimiento=id)
    establecimiento.estado = "ACTIVO"
    establecimiento.save()
    return redirect('listar_establecimientos')

# ---------------------- AREA ----------------------

# CREATE
def crear_area(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        id_establecimiento = request.POST.get('establecimiento')

        establecimiento = Establecimiento.objects.get(idEstablecimiento=id_establecimiento)
        Area.objects.create(areaNombre=nombre, idEstablecimiento=establecimiento)

        return redirect('listar_areas')

    establecimientos = Establecimiento.objects.filter(estado="ACTIVO")
    return render(request, 'area_list.html', {'establecimientos': establecimientos})


# READ
def listar_areas(request):
    areas = Area.objects.all()
    establecimientos = Establecimiento.objects.filter(estado="ACTIVO")
    return render(request, 'area_list.html', {'areas': areas, 'establecimientos': establecimientos})

# UPDATE
def editar_area(request, id):
    area = get_object_or_404(Area, idArea=id)
    if request.method == 'POST':
        area.areaNombre = request.POST.get('nombre')
        area.idEstablecimiento_id = request.POST.get('establecimiento')
        area.save()
        return redirect('listar_areas')
    establecimientos = Establecimiento.objects.filter(estado="ACTIVO")
    return render(request, 'area_list.html', {'area': area, 'establecimientos': establecimientos})

# DELETE 
def eliminar_area(request, id):
    area = get_object_or_404(Area, idArea=id)
    area.estado = "INACTIVO"
    area.save()
    return redirect('listar_areas')

def reactivar_area(request, id):
    area = get_object_or_404(Area, idArea=id)
    area.estado = "ACTIVO"
    area.save()
    return redirect('listar_areas')

# ---------------------- SUB-AREA ----------------------

# CREATE
def crear_subarea(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        id_area = request.POST.get('area')
        SubArea.objects.create(subareaNombre=nombre, idArea_id=id_area)
        return redirect('listar_subareas')
    areas = Area.objects.filter(estado="ACTIVO")
    return render(request, 'subarea_list.html', {'areas': areas})

# READ
def listar_subareas(request):
    subareas = SubArea.objects.all()
    areas = Area.objects.filter(estado="ACTIVO")
    return render(request, 'subarea_list.html', {
        'subareas': subareas,
        'areas': areas
    })

# UPDATE
def editar_subarea(request, id):
    subarea = get_object_or_404(SubArea, idSubarea=id)
    if request.method == 'POST':
        subarea.subareaNombre = request.POST.get('nombre')
        subarea.idArea_id = request.POST.get('area')
        subarea.save()
        return redirect('listar_subareas')
    areas = Area.objects.filter(estado="ACTIVO")
    return render(request, 'subarea_list.html', {'subarea': subarea, 'areas': areas})

# DELETE
def eliminar_subarea(request, id):
    subarea = get_object_or_404(SubArea, idSubarea=id)
    subarea.estado = "INACTIVO"
    subarea.save()
    return redirect('listar_subareas')

def reactivar_subarea(request, id):
    subarea = get_object_or_404(SubArea, idSubarea=id)
    subarea.estado = "ACTIVO"
    subarea.save()
    return redirect('listar_subareas')

# ---------------------- CARGO ----------------------
# CREATE
def crear_cargo(request):
    if request.method == 'POST':
        descripcion = request.POST.get('descripcion')
        Cargo.objects.create(Descripcion=descripcion)
        return redirect('listar_cargos')
    return render(request, 'cargo_list.html')

# READ
def listar_cargos(request):
    cargos = Cargo.objects.all()
    return render(request, 'cargo_list.html', {'cargos': cargos})

# UPDATE
def editar_cargo(request, id):
    cargo = get_object_or_404(Cargo, idCargo=id)
    if request.method == 'POST':
        cargo.Descripcion = request.POST.get('descripcion')
        cargo.save()
        return redirect('listar_cargos')
    return render(request, 'cargo_list.html', {'cargo': cargo})

# DELETE 
def eliminar_cargo(request, id):
    cargo = get_object_or_404(Cargo, idCargo=id)
    cargo.estado = "INACTIVO"
    cargo.save()
    return redirect('listar_cargos')

def reactivar_cargo(request, id):
    cargo = get_object_or_404(Cargo, idCargo=id)
    cargo.estado = "ACTIVO"
    cargo.save()
    return redirect('listar_cargos')

# ---------------------- PERSONAL ----------------------

def crear_personal(request):
    if request.method == 'POST':
        dni = request.POST.get('dni')
        id_cargo = request.POST.get('cargo')
        id_area = request.POST.get('area')
        apellidos = request.POST.get('apellidos')
        nombres = request.POST.get('nombres')
        telefono = request.POST.get('telefono')
        correo = request.POST.get('correo')

        password = request.POST.get('password')
        username = request.POST.get('username')

        user = None
        if username and password:
            user = User.objects.create(
                username=username,
                first_name=nombres,
                last_name=apellidos,
                email=correo,
                is_superuser=False,
                is_staff=False,
                is_active="ACTIVO",
                password=make_password(password)
            )

        Personal.objects.create(
            dni=dni,
            idCargo_id=id_cargo,
            idArea_id=id_area,
            Apellidos=apellidos,
            Nombres=nombres,
            Telefono=telefono,
            Correo=correo,
            user=user
        )

        return redirect('listar_personal')

    cargos = Cargo.objects.all()
    establecimientos = Establecimiento.objects.filter(estado="ACTIVO")
    return render(request, 'personal_list.html', {'cargos': cargos, 'establecimientos': establecimientos})

def asignar_usuario(request):
    if request.method == 'POST':
        id_personal = request.POST.get('idPersonal')
        username = request.POST.get('username')
        password = request.POST.get('password')

        personal = get_object_or_404(Personal, idPersonal=id_personal)

        if not personal.user:
            user = User.objects.create(
                username=username,
                first_name=personal.Nombres,
                last_name=personal.Apellidos,
                email=personal.Correo,
                password=make_password(password),
                is_active=True
            )
            personal.user = user
            personal.save()

        return redirect('listar_personal')

# READ
def listar_personal(request):
    personal = Personal.objects.all()
    cargos = Cargo.objects.all()
    areas = Area.objects.filter(estado="ACTIVO")
    return render(request, 'personal_list.html', {
        'personal': personal,
        'cargos': cargos,
        'areas': areas
    })


# UPDATE
def editar_personal(request, id):
    persona = get_object_or_404(Personal, idPersonal=id)

    if request.method == 'POST':
        dni = request.POST.get('dni')
        id_cargo = request.POST.get('cargo')
        id_area = request.POST.get('area')
        apellidos = request.POST.get('apellidos')
        nombres = request.POST.get('nombres')
        telefono = request.POST.get('telefono')
        correo = request.POST.get('correo')


        persona.dni = dni
        persona.idCargo_id = id_cargo
        persona.idArea_id = id_area
        persona.Apellidos = apellidos
        persona.Nombres = nombres
        persona.Telefono = telefono
        persona.Correo = correo
        persona.save()

        return redirect('listar_personal')

    cargos = Cargo.objects.all()
    establecimientos = Establecimiento.objects.filter(estado="ACTIVO")

    return render(request, 'personal_list.html', {
        'persona': persona,
        'cargos': cargos,
        'establecimientos': establecimientos
    })

# DELETE 
def eliminar_personal(request, id):
    persona = get_object_or_404(Personal, idPersonal=id)
    user = persona.user 
    persona.estado = "INACTIVO"
    persona.save()
    return redirect('listar_personal')

def reactivar_personal(request, id):
    persona = get_object_or_404(Personal, idPersonal=id)
    persona.estado = "ACTIVO"
    persona.save()
    return redirect('listar_personal')

# ---------------------- FORMATO ----------------------

# CREATE
def crear_formato(request):
    if request.method == 'POST':
        codigo = request.POST.get('codigoFormato')
        version = request.POST.get('versionFormato')
        fecha_str = request.POST.get('fechaFormato') 
        fecha = datetime.strptime(fecha_str, '%Y-%m-%dT%H:%M')
        estado = "ACTIVO"
        Formato.objects.create(
            codigoFormato=codigo,
            versionFormato=version,
            fechaFormato=fecha,
            estadoFormato=estado
        )
        return redirect('listar_formatos')
    return render(request, 'formato_list.html')

# READ
def listar_formatos(request):
    formatos = Formato.objects.all()
    return render(request, 'formato_list.html', {'formatos': formatos})

# UPDATE
def editar_formato(request, id):
    formato = get_object_or_404(Formato, idFormato=id)
    if request.method == 'POST':
        formato.codigoFormato = request.POST.get('codigoFormato')
        formato.versionFormato = request.POST.get('versionFormato')
        fecha_str = request.POST.get('fechaFormato')
        formato.fechaFormato = datetime.strptime(fecha_str, '%Y-%m-%dT%H:%M')
        formato.save()
        return redirect('listar_formatos')
    return render(request, 'formato_list.html', {'formato': formato})


def eliminar_formato(request, id):
    formato = get_object_or_404(Formato, idFormato=id)
    formato.estadoFormato = 'INACTIVO'
    formato.save()
    return redirect('listar_formatos')

def reactivar_formato(request, id):
    formato = get_object_or_404(Formato, idFormato=id)
    formato.estadoFormato = 'ACTIVO'
    formato.save()
    return redirect('listar_formatos')

# ---------------------- ACTAS SALIDAS ----------------------

def crear_actasalida(request):
    if request.method == 'POST':
        codreq = request.POST.get('codReq')
        formato_id = request.POST.get('idFormato')
        establecimiento_id = request.POST.get('idEstablecimiento')
        personal_id = request.POST.get('idPersonal')

        ActaSalida.objects.create(
            codReq=codreq,
            idFormato_id=formato_id,
            idEstablecimiento_id=establecimiento_id,
            idPersonal_id=personal_id
        )
        return redirect('listar_actasalida')

    formatos = Formato.objects.filter(estadoFormato='ACTIVO').first()
    establecimientos = Establecimiento.objects.filter(estado="ACTIVO")
    personal = Personal.objects.all()

    context = {
        'formatos': formatos,
        'establecimientos': establecimientos,
        'personal': personal,
    }
    return render(request, 'acta_list.html', context)

# READ
def listar_actasalida(request):
    actas = ActaSalida.objects.all()
    formatos = Formato.objects.filter(estadoFormato='ACTIVO').first()
    establecimientos = Establecimiento.objects.filter(estado="ACTIVO")
    personal = Personal.objects.all()
    equipos = EquiposInformatico.objects.all()
    return render(request, 'acta_list.html', {'actas': actas, 'formatos': formatos,
        'establecimientos': establecimientos,
        'personal': personal,
        'equipos': equipos })

# UPDATE
def editar_actasalida(request, id):
    acta = get_object_or_404(ActaSalida, idActaSalida=id)
    if request.method == 'POST':
        acta.codReq = request.POST.get('codReq')
        acta.idFormato_id = request.POST.get('idFormato')
        acta.idEstablecimiento_id = request.POST.get('idEstablecimiento')
        acta.idPersonal_id = request.POST.get('idPersonal')
        acta.idEquipoInformatico_id = request.POST.get('idEquipoInformatico')
        acta.save()
        return redirect('listar_actasalida')
    formatos = Formato.objects.filter(estadoFormato__iexact='activo')
    establecimientos = Establecimiento.objects.filter(estado="ACTIVO")
    personal = Personal.objects.all()
    equipos = EquiposInformatico.objects.all()
    context = {
        'acta': acta,
        'formatos': formatos,
        'establecimientos': establecimientos,
        'personal': personal,
        'equipos': equipos
    }
    return render(request, 'acta_list.html', context)

# DELETE
def eliminar_actasalida(request, id):
    acta = get_object_or_404(ActaSalida, idActaSalida=id)
    acta.estado = "INACTIVO"
    acta.save()
    return redirect('listar_actasalida')

def reactivar_actasalida(request, id):
    acta = get_object_or_404(ActaSalida, idActaSalida=id)
    acta.estado = "ACTIVO"
    acta.save()
    return redirect('listar_actasalida')

def asignar_equipos(request, id):
    acta = get_object_or_404(ActaSalida, idActaSalida=id)

    if request.method == 'POST':
        equipo_id = request.POST.get('idEquipoInformatico')

        if not equipo_id:
            messages.error(request, "Debe seleccionar un equipo.")
            return redirect('asignar_equipo_acta', id=acta.idActaSalida)

        equipo = get_object_or_404(EquiposInformatico, pk=equipo_id)

        if AsignacionEquipo.objects.filter(idEquipoInformatico=equipo).exists():
            messages.error(request, "Este equipo ya está asignado.")
            return redirect('asignar_equipo_acta', id=acta.idActaSalida)

        AsignacionEquipo.objects.create(idActaSalida=acta, idEquipoInformatico=equipo)
        equipo.estado = 'ASIGNADO'
        equipo.save()

        messages.success(request, "✅ Equipo asignado correctamente.")
        return redirect('asignar_equipo_acta', id=acta.idActaSalida)

    buscar_disponibles = request.GET.get('buscar_disponibles', '').strip()
    buscar_asignados = request.GET.get('buscar_asignados', '').strip()

    equipos_qs = EquiposInformatico.objects.filter(estado='ACTIVO')
    if buscar_disponibles:
        equipos_qs = equipos_qs.filter(
            Q(idTipoEquipo__nombre__icontains=buscar_disponibles) |
            Q(numeroSerie__icontains=buscar_disponibles) |
            Q(descripcionEquipo__icontains=buscar_disponibles) |
            Q(idModelo__nombreModelo__icontains=buscar_disponibles) |
            Q(codigoPatrimonial__icontains=buscar_disponibles) |
            Q(idModelo__idMarca__nombreMarca__icontains=buscar_disponibles)
        )

    asignados_qs = AsignacionEquipo.objects.filter(idActaSalida=acta)
    if buscar_asignados:
        asignados_qs = asignados_qs.filter(
            Q(idEquipoInformatico__idTipoEquipo__nombre__icontains=buscar_asignados) |
            Q(idEquipoInformatico__numeroSerie__icontains=buscar_asignados) |
            Q(idEquipoInformatico__descripcionEquipo__icontains=buscar_asignados) |
            Q(idEquipoInformatico__idModelo__idMarca__nombreMarca__icontains=buscar_asignados) |
            Q(idEquipoInformatico__idModelo__nombreModelo__icontains=buscar_asignados)
        )

    # Paginación
    paginador_disponibles = Paginator(equipos_qs, 10)
    paginador_asignados = Paginator(asignados_qs, 10)

    pagina_disponibles = request.GET.get('pagina_disponibles')
    pagina_asignados = request.GET.get('pagina_asignados')

    equipos = paginador_disponibles.get_page(pagina_disponibles)
    equipos_asignados = paginador_asignados.get_page(pagina_asignados)

    context = {
        'acta': acta,
        'equipos': equipos,
        'equipos_asignados': equipos_asignados,
        'buscar_disponibles': buscar_disponibles,
        'buscar_asignados': buscar_asignados,
    }
    return render(request, 'asignacion_equipos.html', context)


def liberar_equipo_acta(request, id_acta, id_asignacion):
    asignacion = get_object_or_404(AsignacionEquipo, pk=id_asignacion, idActaSalida__idActaSalida=id_acta)
    equipo = asignacion.idEquipoInformatico

    if request.method == 'POST':
        equipo.estado = "ACTIVO"
        equipo.save()

        asignacion.delete()

        messages.success(request, "✅ Equipo liberado y marcado como DISPONIBLE.")

    return redirect('asignar_equipo_acta', id=id_acta)

def detalle_acta_salida(request, id):
    acta = get_object_or_404(ActaSalida, pk=id)
    equipos_asignados = AsignacionEquipo.objects.filter(idActaSalida=acta)
    personal = Personal.objects.all()
    establecimiento = Establecimiento.objects.all()

    contexto = {
        'acta': acta,
        'personales': personal,
        'establecimientos': establecimiento,  
        'equipos_asignados': equipos_asignados,
    }
    return render(request, 'acta_reporte.html', contexto)
# ---------------------- LOTE ----------------------
# CREATE
def crear_lote(request):
    if request.method == 'POST':
        cantidad = request.POST.get('cantidad')
        modelo_id = request.POST.get('idModelo')
        fecha_adq = request.POST.get('fechaAdquisicion')

        lote = Lote.objects.create(
            cantidad=cantidad,
            idModelo_id=modelo_id,
            fechaAdquisicion=fecha_adq
        )

        return JsonResponse({'idLote': lote.idLote})

    modelos = Modelo.objects.all()
    return render(request, 'equipo_list.html', {'modelos': modelos})

# READ
def listar_lotes(request):
    lotes = Lote.objects.all()
    return render(request, 'equipo_list.html', {'lotes': lotes})

# UPDATE
def editar_lote(request, id):
    lote = get_object_or_404(Lote, idLote=id)
    if request.method == 'POST':
        lote.cantidad = request.POST.get('cantidad')
        lote.idModelo_id = request.POST.get('idModelo')
        lote.fechaAdquisicion = request.POST.get('fechaAdquisicion')
        lote.save()
        return redirect('listar_lotes')

    modelos = Modelo.objects.all()
    return render(request, 'equipo_list.html', {'lote': lote, 'modelos': modelos})

# DELETE
def eliminar_lote(request, id):
    lote = get_object_or_404(Lote, idLote=id)
    lote.estado = "INACTIVO"
    lote.save()
    return redirect('listar_lotes')

def reactivar_lote(request, id):
    lote = get_object_or_404(Lote, idLote=id)
    lote.estado = "ACTIVO"
    lote.save()
    return redirect('listar_lotes')

# ---------------------- USUARIOS ----------------------

def listar_usuarios(request):
    personal = Personal.objects.select_related('user', 'idArea', 'idCargo') \
        .filter(idArea__areaNombre="OGI") \
        .order_by('Nombres')

    return render(request, 'usuario_list.html', {'personal': personal})


def historial_equipos(request):
    estado = request.GET.get('estado', '').upper()
    tipo = request.GET.get('tipo', '')
    buscar = request.GET.get('buscar', '')

    equipos = EquiposInformatico.objects.select_related('idTipoEquipo', 'idModelo', 'idModelo__idMarca')

    # Filtro por estado
    if estado in ["ACTIVO", "INACTIVO", "MANTENIMIENTO", "BAJA"]:
        equipos = equipos.filter(estado=estado)

    # Filtro por tipo
    if tipo:
        equipos = equipos.filter(idTipoEquipo__nombre__iexact=tipo)

    # Búsqueda general
    if buscar:
        equipos = equipos.filter(
            Q(nombreEquipoInformatico__icontains=buscar) |
            Q(codigoPatrimonial__icontains=buscar) |
            Q(numeroSerie__icontains=buscar) |
            Q(idModelo__nombreModelo__icontains=buscar) |
            Q(idModelo__idMarca__nombreMarca__icontains=buscar) |
            Q(idTipoEquipo__nombre__icontains=buscar) |
            Q(observacionEquipo__icontains=buscar) |
            Q(descripcionEquipo__icontains=buscar)
        )

    tipos = TipoEquipo.objects.all()
    modelos = Modelo.objects.all()

    return render(request, 'historial_equipos.html', {
        'historial': equipos,
        'tipos': tipos,
        'modelos': modelos,
        'estado_actual': estado.lower(),
        'tipo_actual': tipo.lower(),
    })


def historial_componentes(request):
    estado = request.GET.get('estado', '').upper()
    tipo = request.GET.get('tipo', '')
    buscar = request.GET.get('buscar', '').strip()

    componentes = Componentes.objects.select_related('idModelo', 'idTipoComponente', 'idModelo__idMarca')

    if estado in ["ACTIVO", "INACTIVO"]:
        componentes = componentes.filter(estado=estado)

    if tipo:
        componentes = componentes.filter(idTipoComponente__nombre__iexact=tipo)

    if buscar:
        componentes = componentes.filter(
            Q(nombreComponente__icontains=buscar) |
            Q(idTipoComponente__nombre__icontains=buscar) |
            Q(idModelo__nombreModelo__icontains=buscar) |
            Q(idModelo__idMarca__nombreMarca__icontains=buscar) |
            Q(numeroSerie__icontains=buscar)
        )

    tipos = TipoComponente.objects.all()
    modelos = Modelo.objects.all()

    return render(request, 'historial_componentes.html', {
        'componentes': componentes,
        'tipos': tipos,
        'modelos': modelos,
        'estado_actual': estado.lower(),
        'tipo_actual': tipo.lower(),
    })


def detalle_equipo_pdf(request, idEquipoInformatico):
    equipo = get_object_or_404(EquiposInformatico, idEquipoInformatico=idEquipoInformatico)
    componentes = Componentes.objects.filter(idEquipoInformatico=equipo)

    return generar_pdf(
        nombre_template='estructura_detalle_equipo_pdf.html',
        contexto={
            'equipo': equipo,
            'componentes': componentes
        },
        nombre_archivo=f'detalle_equipo_{idEquipoInformatico}.pdf'
    )

def componente_detalle_pdf(request, idComponente):
    componente = get_object_or_404(Componentes, pk=idComponente)
    caracteristicas = ComponenteCaracteristica.objects.filter(idComponente=componente)

    return generar_pdf(
        nombre_template='estructura_detalle_componente_pdf.html',
        contexto={
            'componente': componente,
            'caracteristicas': caracteristicas
        },
        nombre_archivo=f'detalle_componente_{idComponente}.pdf'
    )

def historial_equipos_pdf(request):
    estado = request.GET.get('estado', '').upper()
    tipo = request.GET.get('tipo', '')

    equipos = EquiposInformatico.objects.select_related('idTipoEquipo', 'idModelo', 'idModelo__idMarca')

    if estado in ["ACTIVO", "INACTIVO"]:
        equipos = equipos.filter(estado=estado)

    if tipo:
        equipos = equipos.filter(idTipoEquipo__nombre__iexact=tipo)

    return generar_pdf(
        nombre_template='historial_equipos_reporte.html',
        contexto={
            'historial': equipos,
        },
        nombre_archivo='reporte_historial_equipos.pdf'
    )

def historial_componentes_pdf(request):
    estado = request.GET.get('estado', '').upper()
    tipo = request.GET.get('tipo', '')

    componentes = Componentes.objects.select_related('idModelo', 'idTipoComponente', 'idModelo__idMarca')

    if estado in ["ACTIVO", "INACTIVO"]:
        componentes = componentes.filter(estado=estado)

    if tipo:
        componentes = componentes.filter(idTipoComponente__nombre__iexact=tipo)

    return generar_pdf(
        nombre_template='historial_componentes_reporte.html',
        contexto={'historial': componentes},
        nombre_archivo='reporte_historial_componentes.pdf'
    )

def exportar_componentes_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Historial de Componentes"

    # Encabezados
    encabezados = ['Item', 'Tipo', 'Marca', 'Modelo', 'N° Serie', 'Estado']
    ws.append(encabezados)

    # Estilo encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Datos
    componentes = Componentes.objects.select_related('idModelo', 'idModelo__idMarca', 'idTipoComponente')

    for i, comp in enumerate(componentes, start=1):
        fila = [
            i,
            comp.idTipoComponente.nombre,
            comp.idModelo.idMarca.nombreMarca,
            comp.idModelo.nombreModelo,
            comp.numeroSerie,
            comp.estado.capitalize(),
        ]
        ws.append(fila)

    # Estilo filas
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=6):
        for cell in row:
            cell.alignment = center_align
            cell.border = thin_border

    # Ajuste automático de columnas
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max_length + 3

    # Exportar como respuesta HTTP
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=historial_componentes.xlsx"
    wb.save(response)
    return response

def exportar_equipos_excel(request):
    estado = request.GET.get('estado', '').upper()
    tipo = request.GET.get('tipo', '')

    equipos = EquiposInformatico.objects.select_related('idTipoEquipo', 'idModelo', 'idModelo__idMarca')

    if estado in ["ACTIVO", "INACTIVO"]:
        equipos = equipos.filter(estado=estado)

    if tipo:
        equipos = equipos.filter(idTipoEquipo__nombre__iexact=tipo)

    wb = Workbook()
    ws = wb.active
    ws.title = "Historial de Equipos"

    # Encabezados
    headers = ['#', 'Patrimonio', 'Tipo', 'Nombre', 'N° Serie', 'Marca', 'Modelo', 'Estado']
    ws.append(headers)

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    # Filas de datos
    for i, equipo in enumerate(equipos, start=1):
        row = [
            i,
            equipo.codigoPatrimonial,
            equipo.idTipoEquipo.nombre,
            equipo.nombreEquipoInformatico,
            equipo.numeroSerie,
            equipo.idModelo.idMarca.nombreMarca,
            equipo.idModelo.nombreModelo,
            equipo.estado.capitalize(),
        ]
        ws.append(row)

    # Aplicar estilo a filas
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=8):
        for cell in row:
            cell.alignment = align_center
            cell.border = thin_border

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 3

    # Respuesta HTTP
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=historial_equipos.xlsx"
    wb.save(response)
    return response

def generar_acta_pdf(request, id_acta):
    from .models import ActaSalida, AsignacionEquipo  # Asegúrate que esté tu import correcto

    acta = ActaSalida.objects.select_related('idFormato').get(idActa=id_acta)
    equipos_asignados = AsignacionEquipo.objects.filter(idActa=acta).select_related(
        'idEquipoInformatico__idModelo__idMarca',
        'idEquipoInformatico__idTipoEquipo'
    )
